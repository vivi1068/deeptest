# -*- coding: utf-8 -*-
# __author__ = 'Carina'


from openpyxl import Workbook
from openpyxl import load_workbook
import datetime


if __name__ == "__main__":
    print("写Excel简单示例")

    # 创建一个excel工作区
    wb = Workbook()

    # 激活当前工作簿
    ws = wb.active

    # 往单元格A1写入数据，其他单元格写入类似
    ws['A1'] = "语言"
    ws['A2'] = 'python'
    ws['B1'] = "时间"

    # 写入时间类型到excel, python会自动将类型转换成excel的日期时间类型
    ws['B2'] = datetime.datetime.now()

    # 写下一行数据，列表元素对应每一个单元格
    ws.append([1, 2, 3])

    # 保存为excel文件
    wb.save("简单excel写示例.xlsx")


    # 读取已存在的excel文件
    print("读取已存在的excel文件")

    wb = load_workbook("简单Excel写示例.xlsx")

    # 获取所有sheet名，返回的是list类型
    sheets = wb.get_sheet_names()
    print(type(sheets))

    # 遍历sheets，并读取其单元格内容打印输出
    for sh in sheets:
        print("读取工作簿名称：", sh)

    # 获取要读取的sheet
    ws = wb.get_sheet_by_name(sheets[0])

    # 读取单个单元格的值
    A1 = ws['A1'].value
    print("A1单元格的值：", A1)

    # 读取A3、B3、C3
    for index in ('A2', 'B2', 'C2'):
        print(index, "单元格的值：", ws[index].value)

    # 读取空单元格，返回None
    C1 = ws['C1'].value
    print("C1单元格的值：", C1)



